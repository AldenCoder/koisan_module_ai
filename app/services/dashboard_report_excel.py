from __future__ import annotations

import json
from datetime import datetime
from enum import Enum
from typing import Any, Callable, Dict, Iterable, List, Sequence, Tuple
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.api.dependencies.time import VN_TZ, now_vn


EXCEL_MAX_CELL_LENGTH = 32767
DASHBOARD_HISTORY_BASE_URL = "https://koisan-dashboard.vercel.app/quan-ly-lich-su"
TITLE_ROW = 1
DESCRIPTION_ROW = 2
COUNT_ROW = 3
HEADER_ROW = 5
DATA_START_ROW = 6

NAVY = "17324D"
TEAL = "0F766E"
BLUE = "2563EB"
CYAN = "0891B2"
PURPLE = "7C3AED"
VIOLET = "6D28D9"
GREEN = "059669"
AMBER = "D97706"
RED = "DC2626"
SLATE = "475569"
WHITE = "FFFFFF"
TEXT = "0F172A"
MUTED = "64748B"

THIN_GRAY = Side(style="thin", color="CBD5E1")
BOTTOM_BORDER = Border(bottom=THIN_GRAY)

ROLE_LABELS = {
    "user": "Khách hàng",
    "staff": "Nhân viên",
    "bot": "Agent",
    "system": "Hệ thống",
}
CONTENT_TYPE_LABELS = {
    "text": "Văn bản",
    "image": "Hình ảnh",
    "other": "Khác / trống",
}
THREAD_TYPE_LABELS = {
    "inbox": "Tin nhắn riêng",
    "comment": "Bình luận",
    "INBOX": "Tin nhắn riêng",
    "COMMENT": "Bình luận",
}
STATUS_LABELS = {
    "new": "Mới",
    "confirmed": "Đã xác nhận",
    "handover": "Cần nhân viên hỗ trợ",
    "apilimit": "Giới hạn API",
    "order_pending": "Chờ xử lý đơn hàng",
}
SUPPORT_REASON_LABELS = {
    "handover": "Chuyển cho nhân viên",
    "apilimit": "Giới hạn API",
    "bot_paused": "Agent đang tạm dừng",
    "needs_support": "Cần hỗ trợ",
}


Column = Tuple[str, str, float, str]


MESSAGE_COLUMNS: Sequence[Column] = (
    ("_index", "STT", 8, "index"),
    ("message_mid", "Mã tin nhắn nền tảng", 25, "text"),
    ("conversation_id", "ID hội thoại hệ thống", 24, "system_id_link"),
    ("content_type", "Loại nội dung", 17, "content_type"),
    ("role", "Người gửi", 15, "role"),
    ("content", "Nội dung tin nhắn", 60, "long_text"),
    ("image_urls", "URL hình ảnh", 48, "multiline"),
    ("page_id", "ID trang", 20, "text"),
    ("thread_type", "Loại hội thoại", 18, "thread_type"),
    ("source", "Nguồn dữ liệu", 28, "text"),
    ("sender_id", "ID người gửi", 22, "text"),
    ("sender_name", "Tên người gửi", 24, "text"),
    ("pancake_conversation_id", "Mã hội thoại nền tảng", 26, "text"),
    ("created_at", "Thời gian tạo", 21, "datetime"),
    ("updated_at", "Thời gian cập nhật", 21, "datetime"),
    ("meta", "Dữ liệu bổ sung (JSON)", 60, "json"),
)

CONVERSATION_COLUMNS: Sequence[Column] = (
    ("_index", "STT", 8, "index"),
    ("conversation_id", "ID hội thoại hệ thống", 24, "system_id_link"),
    ("customer_name", "Tên khách hàng", 24, "text"),
    ("customer_id", "ID khách hàng", 22, "text"),
    ("channel", "Kênh / Trang", 24, "text"),
    ("pancake_page_id", "ID trang", 20, "text"),
    ("pancake_conversation_id", "Mã hội thoại nền tảng", 26, "text"),
    ("pancake_thread_type", "Loại hội thoại", 18, "thread_type"),
    ("status", "Trạng thái", 25, "status"),
    ("is_active", "Đang hoạt động", 16, "boolean"),
    ("pancake_info_url", "Liên kết Pancake", 42, "url"),
    ("order_note", "Ghi chú đơn hàng", 52, "long_text"),
    ("summaries", "Tóm tắt hội thoại", 52, "json"),
    ("message_count", "Số tin nhắn", 15, "integer"),
    ("fb_ai_initialized", "Đã khởi tạo AI", 16, "boolean"),
    ("fb_ai_initialized_at", "Thời gian khởi tạo AI", 22, "datetime"),
    ("bot_paused_until", "Agent tạm dừng đến", 22, "datetime"),
    ("bot_paused_at", "Thời gian tạm dừng", 22, "datetime"),
    ("bot_paused_reason", "Lý do tạm dừng", 24, "text"),
    ("bot_paused_by", "Người tạm dừng", 22, "text"),
    ("created_at", "Thời gian tạo", 21, "datetime"),
    ("updated_at", "Thời gian cập nhật", 21, "datetime"),
)

SUPPORT_COLUMNS: Sequence[Column] = (
    CONVERSATION_COLUMNS[0],
    CONVERSATION_COLUMNS[1],
    CONVERSATION_COLUMNS[2],
    CONVERSATION_COLUMNS[3],
    ("support_reason", "Lý do cần hỗ trợ", 25, "support_reason"),
    *CONVERSATION_COLUMNS[4:],
)


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Enum):
        return str(value.value)
    return str(value)


def _json_text(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, default=str)


def _excel_datetime(value: Any) -> Any:
    if not isinstance(value, datetime):
        return value
    if value.tzinfo is None:
        return value
    return value.astimezone(VN_TZ).replace(tzinfo=None)


def _limit_excel_text(value: str) -> str:
    if len(value) <= EXCEL_MAX_CELL_LENGTH:
        return value
    suffix = "\n[Đã rút gọn do giới hạn 32.767 ký tự của một ô Excel]"
    return value[: EXCEL_MAX_CELL_LENGTH - len(suffix)] + suffix


def _format_value(value: Any, kind: str, *, index: int) -> Any:
    if kind == "index":
        return index
    if kind == "datetime":
        return _excel_datetime(value)
    if kind == "boolean":
        return "Có" if bool(value) else "Không"
    if kind == "integer":
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0
    if kind == "json":
        return _limit_excel_text(_json_text(value))

    text = _as_text(value)
    if kind == "role":
        return ROLE_LABELS.get(text.lower(), text or "Không xác định")
    if kind == "content_type":
        return CONTENT_TYPE_LABELS.get(text.lower(), text or "Không xác định")
    if kind == "thread_type":
        return THREAD_TYPE_LABELS.get(text, THREAD_TYPE_LABELS.get(text.lower(), text))
    if kind == "status":
        return STATUS_LABELS.get(text.lower(), text or "Không xác định")
    if kind == "support_reason":
        return SUPPORT_REASON_LABELS.get(text.lower(), text or "Cần hỗ trợ")
    return _limit_excel_text(text)


def _set_cell_value(cell: Cell, value: Any) -> None:
    cell.value = value
    # Nội dung người dùng bắt đầu bằng "=" phải là chuỗi, không phải công thức Excel.
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def _system_id_url(value: Any) -> str:
    system_id = _as_text(value).strip()
    if not system_id:
        return ""
    return f"{DASHBOARD_HISTORY_BASE_URL}/{quote(system_id, safe='')}"


def _style_sheet_title(
    sheet: Any,
    *,
    title: str,
    description: str,
    record_count: int,
    last_column: int,
    accent: str,
) -> None:
    last_letter = get_column_letter(last_column)
    sheet.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=last_column)
    sheet.merge_cells(
        start_row=DESCRIPTION_ROW,
        start_column=1,
        end_row=DESCRIPTION_ROW,
        end_column=last_column,
    )
    sheet.merge_cells(start_row=COUNT_ROW, start_column=1, end_row=COUNT_ROW, end_column=last_column)

    title_cell = sheet.cell(TITLE_ROW, 1, title)
    title_cell.font = Font(name="Arial", size=18, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center")

    description_cell = sheet.cell(DESCRIPTION_ROW, 1, description)
    description_cell.font = Font(name="Arial", size=10, color=SLATE)
    description_cell.fill = PatternFill("solid", fgColor="F8FAFC")
    description_cell.alignment = Alignment(vertical="center", wrap_text=True)

    count_cell = sheet.cell(COUNT_ROW, 1, f"Số bản ghi: {record_count:,}")
    count_cell.font = Font(name="Arial", size=11, bold=True, color=accent)
    count_cell.fill = PatternFill("solid", fgColor=WHITE)
    count_cell.alignment = Alignment(vertical="center")

    for row in (TITLE_ROW, DESCRIPTION_ROW, COUNT_ROW):
        for cell in sheet[f"A{row}:{last_letter}{row}"][0]:
            if row == TITLE_ROW:
                cell.fill = PatternFill("solid", fgColor=NAVY)

    sheet.row_dimensions[TITLE_ROW].height = 30
    sheet.row_dimensions[DESCRIPTION_ROW].height = 32
    sheet.row_dimensions[COUNT_ROW].height = 24
    sheet.sheet_properties.tabColor = accent
    sheet.sheet_view.showGridLines = False


def _add_table(sheet: Any, *, last_row: int, last_column: int, table_name: str) -> None:
    if last_row < DATA_START_ROW:
        sheet.auto_filter.ref = (
            f"A{HEADER_ROW}:{get_column_letter(last_column)}{HEADER_ROW}"
        )
        return
    table = Table(
        displayName=table_name,
        ref=f"A{HEADER_ROW}:{get_column_letter(last_column)}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _create_detail_sheet(
    workbook: Workbook,
    *,
    name: str,
    title: str,
    description: str,
    rows: Iterable[Dict[str, Any]],
    columns: Sequence[Column],
    accent: str,
    table_name: str,
) -> Any:
    row_items = list(rows)
    sheet = workbook.create_sheet(name)
    _style_sheet_title(
        sheet,
        title=title,
        description=description,
        record_count=len(row_items),
        last_column=len(columns),
        accent=accent,
    )

    for column_index, (_, header, width, _) in enumerate(columns, start=1):
        cell = sheet.cell(HEADER_ROW, column_index, header)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.row_dimensions[HEADER_ROW].height = 34

    for row_index, item in enumerate(row_items, start=1):
        excel_row = DATA_START_ROW + row_index - 1
        for column_index, (key, _, _, kind) in enumerate(columns, start=1):
            value = _format_value(item.get(key), kind, index=row_index)
            cell = sheet.cell(excel_row, column_index)
            _set_cell_value(cell, value)
            cell.font = Font(name="Arial", size=10, color=TEXT)
            cell.border = BOTTOM_BORDER
            cell.alignment = Alignment(
                horizontal="right" if kind in {"integer", "index"} else "left",
                vertical="top",
                wrap_text=kind in {"long_text", "multiline", "json"},
            )
            if kind == "datetime" and isinstance(value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm:ss"
            elif kind in {"integer", "index"}:
                cell.number_format = "#,##0"
            elif kind == "system_id_link" and value:
                cell.hyperlink = _system_id_url(value)
                cell.style = "Hyperlink"
            elif kind == "url" and isinstance(value, str) and value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.style = "Hyperlink"
        sheet.row_dimensions[excel_row].height = 36

    if not row_items:
        empty_cell = sheet.cell(DATA_START_ROW, 1, "Không có dữ liệu phù hợp với bộ lọc đã chọn.")
        empty_cell.font = Font(name="Arial", size=10, italic=True, color=MUTED)
        empty_cell.alignment = Alignment(vertical="center")
        sheet.merge_cells(
            start_row=DATA_START_ROW,
            start_column=1,
            end_row=DATA_START_ROW,
            end_column=len(columns),
        )
        sheet.row_dimensions[DATA_START_ROW].height = 28

    last_data_row = DATA_START_ROW + len(row_items) - 1
    _add_table(
        sheet,
        last_row=last_data_row,
        last_column=len(columns),
        table_name=table_name,
    )
    sheet.freeze_panes = f"A{DATA_START_ROW}"
    sheet.auto_filter.ref = (
        f"A{HEADER_ROW}:{get_column_letter(len(columns))}{max(HEADER_ROW, last_data_row)}"
    )
    sheet.print_title_rows = f"1:{HEADER_ROW}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    return sheet


def _card(
    sheet: Any,
    *,
    start_column: int,
    start_row: int,
    label: str,
    value: Any,
    color: str,
) -> None:
    end_column = start_column + 1
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row,
        end_column=end_column,
    )
    sheet.merge_cells(
        start_row=start_row + 1,
        start_column=start_column,
        end_row=start_row + 2,
        end_column=end_column,
    )
    label_cell = sheet.cell(start_row, start_column, label)
    value_cell = sheet.cell(start_row + 1, start_column, int(value or 0))
    label_cell.font = Font(name="Arial", size=10, bold=True, color=SLATE)
    value_cell.font = Font(name="Arial", size=20, bold=True, color=color)
    value_cell.number_format = "#,##0"
    for row in range(start_row, start_row + 3):
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
            cell.border = Border(
                left=THIN_GRAY if column == start_column else Side(style=None),
                right=THIN_GRAY if column == end_column else Side(style=None),
                top=THIN_GRAY if row == start_row else Side(style=None),
                bottom=THIN_GRAY if row == start_row + 2 else Side(style=None),
            )
            cell.alignment = Alignment(vertical="center", horizontal="left")


def _filter_display(key: str, value: Any) -> str:
    if key in {"from_date", "to_date"} and isinstance(value, datetime):
        localized = value if value.tzinfo else value.replace(tzinfo=VN_TZ)
        return localized.astimezone(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")
    if key == "thread_type":
        return THREAD_TYPE_LABELS.get(_as_text(value), "Tất cả") if value else "Tất cả"
    if key == "role":
        return ROLE_LABELS.get(_as_text(value).lower(), "Tất cả") if value else "Tất cả"
    if key == "include_inactive":
        return "Có" if bool(value) else "Không"
    return _as_text(value) or "Tất cả"


def _create_summary_sheet(
    workbook: Workbook,
    *,
    report: Dict[str, Any],
    detail_counts: Dict[str, int],
) -> Any:
    sheet = workbook.active
    sheet.title = "Tổng quan"
    sheet.sheet_properties.tabColor = NAVY
    sheet.sheet_view.showGridLines = False
    for column in range(1, 11):
        sheet.column_dimensions[get_column_letter(column)].width = 16

    sheet.merge_cells("A1:J1")
    sheet["A1"] = "BÁO CÁO DỮ LIỆU DASHBOARD"
    sheet["A1"].font = Font(name="Arial", size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    for cell in sheet["A1:J1"][0]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:J2")
    sheet["A2"] = "File đối soát chi tiết, mỗi chỉ số có tab chi tiết tương ứng."
    sheet["A2"].font = Font(name="Arial", size=10, color=SLATE)
    sheet["A2"].fill = PatternFill("solid", fgColor="F8FAFC")
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 36

    summary = report.get("summary") or {}
    cards = [
        ("Tổng tin nhắn", "total_messages", CYAN),
        ("Tin nhắn văn bản", "text_messages", BLUE),
        ("Tin nhắn hình ảnh", "image_messages", PURPLE),
        ("Tin nhắn khách", "user_messages", GREEN),
        ("Tin nhắn nhân viên", "staff_messages", VIOLET),
        ("Tin nhắn Agent", "bot_messages", CYAN),
        ("Hội thoại tạo mới", "total_conversations", BLUE),
        ("Cần hỗ trợ", "needs_support_count", RED),
        ("Cảnh báo đơn hàng", "order_alert_count", AMBER),
    ]
    for index, (label, key, color) in enumerate(cards[:5]):
        _card(
            sheet,
            start_column=1 + index * 2,
            start_row=4,
            label=label,
            value=summary.get(key),
            color=color,
        )
    for index, (label, key, color) in enumerate(cards[5:]):
        _card(
            sheet,
            start_column=1 + index * 2,
            start_row=8,
            label=label,
            value=summary.get(key),
            color=color,
        )

    sheet.merge_cells("I8:J8")
    sheet.merge_cells("I9:J10")
    sheet["I8"] = "Thời điểm xuất file"
    exported_at = report.get("exported_at") or now_vn()
    exported_at = _excel_datetime(exported_at)
    sheet["I9"] = exported_at
    sheet["I9"].number_format = "dd/mm/yyyy hh:mm:ss"
    for row in range(8, 11):
        for column in range(9, 11):
            cell = sheet.cell(row, column)
            cell.fill = PatternFill("solid", fgColor="F8FAFC")
            cell.border = Border(
                left=THIN_GRAY if column == 9 else Side(style=None),
                right=THIN_GRAY if column == 10 else Side(style=None),
                top=THIN_GRAY if row == 8 else Side(style=None),
                bottom=THIN_GRAY if row == 10 else Side(style=None),
            )
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    sheet["I8"].font = Font(name="Arial", size=10, bold=True, color=SLATE)
    sheet["I9"].font = Font(name="Arial", size=13, bold=True, color=NAVY)

    sheet.merge_cells("A13:D13")
    sheet["A13"] = "BỘ LỌC ĐÃ ÁP DỤNG"
    sheet["A13"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
    sheet["A13"].fill = PatternFill("solid", fgColor=SLATE)
    for cell in sheet["A13:D13"][0]:
        cell.fill = PatternFill("solid", fgColor=SLATE)

    filter_labels = [
        ("from_date", "Từ ngày"),
        ("to_date", "Đến ngày"),
        ("page_id", "Trang"),
        ("thread_type", "Loại hội thoại"),
        ("role", "Người gửi"),
    ]
    filters = report.get("filters") or {}
    for offset, (key, label) in enumerate(filter_labels, start=14):
        sheet.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=4)
        sheet.cell(offset, 1, label)
        sheet.cell(offset, 2, _filter_display(key, filters.get(key)))
        sheet.cell(offset, 1).font = Font(name="Arial", size=10, bold=True, color=SLATE)
        sheet.cell(offset, 2).font = Font(name="Arial", size=10, color=TEXT)
        for column in range(1, 5):
            cell = sheet.cell(offset, column)
            cell.border = BOTTOM_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[offset].height = 30

    sheet.merge_cells("F13:J13")
    sheet["F13"] = "ĐỐI SOÁT SỐ LIỆU"
    sheet["F13"].font = Font(name="Arial", size=11, bold=True, color=WHITE)
    sheet["F13"].fill = PatternFill("solid", fgColor=TEAL)
    for cell in sheet["F13:J13"][0]:
        cell.fill = PatternFill("solid", fgColor=TEAL)

    headers = ["Chỉ số", "Dashboard", "Dòng chi tiết", "Kết quả", "Tab dữ liệu"]
    for column, header in enumerate(headers, start=6):
        cell = sheet.cell(14, column, header)
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SLATE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    reconciliation = [
        ("Tổng tin nhắn", "total_messages", "all_messages", "Tất cả tin nhắn"),
        ("Tin nhắn văn bản", "text_messages", "text_messages", "Tin nhắn văn bản"),
        ("Tin nhắn hình ảnh", "image_messages", "image_messages", "Tin nhắn hình ảnh"),
        ("Tin nhắn khách", "user_messages", "user_messages", "Tin nhắn khách"),
        ("Tin nhắn nhân viên", "staff_messages", "staff_messages", "Tin nhắn nhân viên"),
        ("Tin nhắn Agent", "bot_messages", "bot_messages", "Tin nhắn Agent"),
        ("Hội thoại tạo mới", "total_conversations", "new_conversations", "Hội thoại tạo mới"),
        ("Cần hỗ trợ", "needs_support_count", "needs_support", "Cần hỗ trợ"),
        ("Cảnh báo đơn hàng", "order_alert_count", "orders", "Cảnh báo đơn hàng"),
    ]
    for row, (label, summary_key, detail_key, tab_name) in enumerate(reconciliation, start=15):
        dashboard_count = int(summary.get(summary_key) or 0)
        detail_count = int(detail_counts.get(detail_key) or 0)
        values = [
            label,
            dashboard_count,
            detail_count,
            "ĐỦ" if dashboard_count == detail_count else "CHÊNH LỆCH",
            tab_name,
        ]
        for column, value in enumerate(values, start=6):
            cell = sheet.cell(row, column, value)
            cell.font = Font(name="Arial", size=9, color=TEXT)
            cell.border = BOTTOM_BORDER
            cell.alignment = Alignment(
                horizontal="right" if column in {7, 8} else "left",
                vertical="center",
                wrap_text=True,
            )
            if column in {7, 8}:
                cell.number_format = "#,##0"
        result_cell = sheet.cell(row, 9)
        result_cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        result_cell.fill = PatternFill(
            "solid",
            fgColor=GREEN if dashboard_count == detail_count else RED,
        )
        result_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A4"
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    return sheet


def _message_rows(messages: Iterable[Dict[str, Any]], predicate: Callable[[Dict[str, Any]], bool]) -> List[Dict[str, Any]]:
    return [message for message in messages if predicate(message)]


def build_dashboard_report_workbook(report: Dict[str, Any]) -> Workbook:
    details = report.get("export_details") or {}
    all_messages = list(details.get("messages") or [])
    text_messages = _message_rows(
        all_messages,
        lambda row: _as_text(row.get("content_type")).lower() == "text",
    )
    image_messages = _message_rows(
        all_messages,
        lambda row: _as_text(row.get("content_type")).lower() == "image",
    )
    user_messages = _message_rows(
        all_messages,
        lambda row: _as_text(row.get("role")).lower() == "user",
    )
    staff_messages = _message_rows(
        all_messages,
        lambda row: _as_text(row.get("role")).lower() == "staff",
    )
    bot_messages = _message_rows(
        all_messages,
        lambda row: _as_text(row.get("role")).lower() == "bot",
    )
    new_conversations = list(details.get("new_conversations") or [])
    needs_support = list(details.get("needs_support") or [])
    orders = list(details.get("orders") or [])

    detail_counts = {
        "all_messages": len(all_messages),
        "text_messages": len(text_messages),
        "image_messages": len(image_messages),
        "user_messages": len(user_messages),
        "staff_messages": len(staff_messages),
        "bot_messages": len(bot_messages),
        "new_conversations": len(new_conversations),
        "needs_support": len(needs_support),
        "orders": len(orders),
    }

    workbook = Workbook()
    _create_summary_sheet(workbook, report=report, detail_counts=detail_counts)

    daily_columns: Sequence[Column] = (
        ("_index", "STT", 8, "index"),
        ("date", "Ngày", 16, "text"),
        ("total", "Tổng tin nhắn", 17, "integer"),
        ("text", "Văn bản", 14, "integer"),
        ("image", "Hình ảnh", 14, "integer"),
        ("user", "Khách hàng", 14, "integer"),
        ("staff", "Nhân viên", 14, "integer"),
        ("bot", "Agent", 14, "integer"),
    )
    _create_detail_sheet(
        workbook,
        name="Theo ngày",
        title="THỐNG KÊ TIN NHẮN THEO NGÀY",
        description="Số lượng tin nhắn theo từng ngày trong phạm vi đã chọn.",
        rows=report.get("messages_by_day") or [],
        columns=daily_columns,
        accent=SLATE,
        table_name="ThongKeTheoNgay",
    )

    status_columns: Sequence[Column] = (
        ("_index", "STT", 8, "index"),
        ("status", "Trạng thái hội thoại", 28, "status"),
        ("count", "Số hội thoại", 18, "integer"),
    )
    _create_detail_sheet(
        workbook,
        name="Trạng thái hội thoại",
        title="TRẠNG THÁI HỘI THOẠI",
        description="Số hội thoại tạo mới trong kỳ, phân theo trạng thái hiện tại.",
        rows=report.get("conversation_status") or [],
        columns=status_columns,
        accent=SLATE,
        table_name="TrangThaiHoiThoai",
    )

    message_sheets = [
        (
            "Tất cả tin nhắn",
            "TOÀN BỘ TIN NHẮN",
            "Danh sách đầy đủ mọi tin nhắn khớp với bộ lọc, gồm cả dữ liệu bổ sung để đối soát.",
            all_messages,
            CYAN,
            "TatCaTinNhan",
        ),
        (
            "Tin nhắn văn bản",
            "TIN NHẮN VĂN BẢN",
            "Danh sách đầy đủ các tin nhắn có nội dung văn bản.",
            text_messages,
            BLUE,
            "TinNhanVanBan",
        ),
        (
            "Tin nhắn hình ảnh",
            "TIN NHẮN HÌNH ẢNH",
            "Danh sách đầy đủ các tin nhắn được nhận diện là URL hình ảnh.",
            image_messages,
            PURPLE,
            "TinNhanHinhAnh",
        ),
        (
            "Tin nhắn khách",
            "TIN NHẮN KHÁCH HÀNG",
            "Danh sách đầy đủ tin nhắn do khách hàng gửi.",
            user_messages,
            GREEN,
            "TinNhanKhach",
        ),
        (
            "Tin nhắn nhân viên",
            "TIN NHẮN NHÂN VIÊN",
            "Danh sách đầy đủ tin nhắn do nhân viên gửi.",
            staff_messages,
            VIOLET,
            "TinNhanNhanVien",
        ),
        (
            "Tin nhắn Agent",
            "TIN NHẮN AGENT",
            "Danh sách đầy đủ tin nhắn do Agent tự động gửi.",
            bot_messages,
            TEAL,
            "TinNhanAgent",
        ),
    ]
    for name, title, description, rows, accent, table_name in message_sheets:
        _create_detail_sheet(
            workbook,
            name=name,
            title=title,
            description=description,
            rows=rows,
            columns=MESSAGE_COLUMNS,
            accent=accent,
            table_name=table_name,
        )

    _create_detail_sheet(
        workbook,
        name="Hội thoại tạo mới",
        title="HỘI THOẠI TẠO MỚI",
        description="Toàn bộ hội thoại được tạo trong khoảng thời gian và bộ lọc đã chọn.",
        rows=new_conversations,
        columns=CONVERSATION_COLUMNS,
        accent=BLUE,
        table_name="HoiThoaiTaoMoi",
    )
    _create_detail_sheet(
        workbook,
        name="Cần hỗ trợ",
        title="HỘI THOẠI CẦN HỖ TRỢ",
        description="Toàn bộ hội thoại cần nhân viên hỗ trợ; không áp dụng giới hạn số dòng của giao diện.",
        rows=needs_support,
        columns=SUPPORT_COLUMNS,
        accent=RED,
        table_name="HoiThoaiCanHoTro",
    )
    _create_detail_sheet(
        workbook,
        name="Cảnh báo đơn hàng",
        title="CẢNH BÁO ĐƠN HÀNG",
        description="Toàn bộ hội thoại có trạng thái chờ đơn hoặc có ghi chú đơn hàng; không giới hạn số dòng.",
        rows=orders,
        columns=CONVERSATION_COLUMNS,
        accent=AMBER,
        table_name="CanhBaoDonHang",
    )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.active = 0
    return workbook
