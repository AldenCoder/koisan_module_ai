from __future__ import annotations

import os
import tempfile
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from zipfile import BadZipFile

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from starlette.concurrency import run_in_threadpool

from app.api.dependencies.error_codes import ErrorCode
from app.api.schemas.product_import import ProductImportResponse

PRODUCT_MARKDOWN_PATH = Path("data/product.md")
PRODUCT_EXCEL_HEADERS = [
    "SKU",
    "Tên",
    "Mô tả",
    "Giá",
    "Ảnh Lookbook",
    "Tồn kho",
    "Loại",
    "Size",
    "Giới tính",
    "Chất liệu",
    "Ảnh Cận chất",
    "Màu",
    "Chiều dài",
    "Phụ kiện tặng kèm",
    "Ảnh feedback",
]
PRODUCT_EXCEL_HEADER_ALIASES = {
    "Code": "SKU",
}


class ProductImportError(Exception):
    def __init__(
        self,
        error_code: ErrorCode,
        status_code: int,
        message: str | None = None,
    ) -> None:
        super().__init__(message or error_code.value)
        self.error_code = error_code
        self.status_code = status_code


def _path_for_response(path: Path) -> str:
    resolved = path.resolve()
    try:
        value = resolved.relative_to(Path.cwd().resolve())
    except ValueError:
        value = resolved
    return str(value).replace("\\", "/")


def _normalize_header_name(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def _normalize_cell_text(value: str) -> str:
    lines = [
        line.strip()
        for line in value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    ]
    lines = [line for line in lines if line]
    if not lines:
        return ""
    return ", ".join(lines)


def _cell_import_value(cell: Any) -> Any:
    hyperlink = getattr(cell, "hyperlink", None)
    target = getattr(hyperlink, "target", None)
    if target:
        return target
    return getattr(cell, "value", None)


def _format_cell_value(value: Any, *, header: str) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value)
    elif isinstance(value, Decimal):
        text = format(value, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
    elif isinstance(value, datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        text = value.isoformat()
    else:
        text = str(value).strip()

    text = _normalize_cell_text(text)
    if header == "SKU":
        return text.upper()
    return text


def _format_markdown_value(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "\n  ")


def _render_products_markdown(
    products: Sequence[dict[str, str]],
    *,
    headers: Sequence[str],
) -> str:
    sections: list[str] = []
    for index, product in enumerate(products, start=1):
        lines = [f"### Sản phẩm {index}", ""]
        for header in headers:
            value = _format_markdown_value(product.get(header, ""))
            lines.append(f"- **{header}**: {value}")
        sections.append("\n".join(lines))
    if not sections:
        return ""
    return "\n\n".join(sections).rstrip() + "\n"


def _write_text_atomically(content: str, destination: Path) -> None:
    temp_path: Path | None = None
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        descriptor, raw_temp_path = tempfile.mkstemp(
            prefix=f".{destination.name}.",
            suffix=".tmp",
            dir=str(destination.parent),
        )
        temp_path = Path(raw_temp_path)
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, destination)
        temp_path = None
    except OSError as exc:
        raise ProductImportError(
            ErrorCode.INTERNAL_SERVER_ERROR,
            500,
            "Failed to write product markdown file",
        ) from exc
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def _load_products_from_excel(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=False,
            data_only=True,
        )
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ProductImportError(
            ErrorCode.INVALID_INPUT_DATA,
            422,
            "Invalid Excel file",
        ) from exc

    try:
        worksheet = workbook.active
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1),
            None,
        )
        if header_row is None:
            raise ProductImportError(
                ErrorCode.INVALID_INPUT_DATA,
                422,
                "Excel file is missing a header row",
            )

        expected_by_normalized = {
            _normalize_header_name(header): header for header in PRODUCT_EXCEL_HEADERS
        }
        expected_by_normalized.update(
            {
                _normalize_header_name(alias): canonical
                for alias, canonical in PRODUCT_EXCEL_HEADER_ALIASES.items()
            }
        )
        column_indexes: dict[str, int] = {}
        for index, raw_header in enumerate(header_row):
            normalized_header = _normalize_header_name(_cell_import_value(raw_header))
            header = expected_by_normalized.get(normalized_header)
            if header is not None and header not in column_indexes:
                column_indexes[header] = index

        products: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=2):
            product: dict[str, str] = {}
            for header in PRODUCT_EXCEL_HEADERS:
                column_index = column_indexes.get(header)
                value = (
                    _cell_import_value(row[column_index])
                    if column_index is not None and column_index < len(row)
                    else None
                )
                product[header] = _format_cell_value(value, header=header)
            if any(product.values()):
                products.append(product)
    finally:
        workbook.close()

    return PRODUCT_EXCEL_HEADERS.copy(), products


def _import_products_content(
    content: bytes,
    *,
    output_path: Path,
) -> ProductImportResponse:
    headers, products = _load_products_from_excel(content)
    markdown = _render_products_markdown(products, headers=headers)
    _write_text_atomically(markdown, output_path)
    return ProductImportResponse(
        output_path=_path_for_response(output_path),
        imported_count=len(products),
        fields=list(headers),
    )


async def import_products_from_excel_service(
    *,
    upload: UploadFile,
    output_path: Path | None = None,
) -> ProductImportResponse:
    filename = upload.filename or ""
    if not filename.lower().endswith(".xlsx"):
        await upload.close()
        raise ProductImportError(
            ErrorCode.INVALID_INPUT_DATA,
            422,
            "Only .xlsx files are supported",
        )

    try:
        content = await upload.read()
        if not content:
            raise ProductImportError(
                ErrorCode.INVALID_INPUT_DATA,
                422,
                "Uploaded Excel file is empty",
            )
        return await run_in_threadpool(
            _import_products_content,
            content,
            output_path=output_path or PRODUCT_MARKDOWN_PATH,
        )
    finally:
        await upload.close()
