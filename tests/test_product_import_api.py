from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.api.dependencies.error_codes import ErrorCode
from app.api.router_v1 import api_router as api_router_v1
from app.api.v1 import products as products_api
from app.services import product_import_service


def _client():
    app = FastAPI()
    app.include_router(products_api.router, prefix="/api/v1/products")
    return TestClient(app)


def _xlsx_bytes(*rows, headers=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers or product_import_service.PRODUCT_EXCEL_HEADERS)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _save_workbook_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_router_v1_registers_product_import_route():
    paths = [route.path for route in api_router_v1.routes]

    assert "/products/import" in paths


def test_product_import_api_writes_markdown_and_overwrites_old_data(
    monkeypatch,
    tmp_path,
):
    output_path = tmp_path / "product.md"
    output_path.write_text("old data", encoding="utf-8")
    monkeypatch.setattr(product_import_service, "PRODUCT_MARKDOWN_PATH", output_path)

    content = _xlsx_bytes(
        [
            "w2651703",
            "Đầm dự tiệc",
            "Đầm dài cổ đổ nhẹ thanh lịch",
            489000,
            "https://example.com/lookbook",
            50,
            "Đầm dự tiệc",
            "",
            "Nữ",
            "Lụa Valentino",
            "",
            "Hồng kem\nHồng\r\nĐỏ đô",
            "130cm",
            "không",
            "https://example.com/feedback",
        ],
        [
            "S2652002",
            "Đầm macxi đi biển",
            "Voan tơ cao cấp",
            369000,
            "",
            "",
            "Đầm macxi đi biển",
            "",
            "Nữ",
            "Voan tơ cao cấp",
            "",
            "Đỏ, Xanh",
            "126cm",
            "không",
            "",
        ],
    )

    response = _client().post(
        "/api/v1/products/import",
        files={
            "file": (
                "products.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    assert response.json()["imported_count"] == 2
    assert response.json()["fields"][0] == "SKU"
    markdown = output_path.read_text(encoding="utf-8")
    assert "old data" not in markdown
    assert "### Sản phẩm 1" in markdown
    assert "- **SKU**: W2651703" in markdown
    assert "- **Tên**: Đầm dự tiệc" in markdown
    assert "- **Giá**: 489000" in markdown
    assert "- **Màu**: Hồng kem, Hồng, Đỏ đô" in markdown
    assert "- **Size**: " in markdown
    assert "shortDescription" not in markdown
    assert "### Sản phẩm 2" in markdown
    assert "- **Tồn kho**: " in markdown


def test_product_import_api_ignores_extra_columns_and_blanks_missing_columns(
    monkeypatch,
    tmp_path,
):
    output_path = tmp_path / "product.md"
    monkeypatch.setattr(product_import_service, "PRODUCT_MARKDOWN_PATH", output_path)
    headers = [
        "Code",
        "Tên",
        "shortDescription",
        "Giá",
        "Extra Column",
        "Ảnh  Lookbook",
        "Chiều dài ",
        "Phụ kiện tặng kèm ",
        "Ảnh Cận chất",
        "Ảnh feedback",
    ]

    content = _xlsx_bytes(
        [
            "s123",
            "Sản phẩm test",
            "Không lưu mô tả ngắn",
            100000,
            "Không lưu cột thừa",
            "https://example.com/lookbook",
            "120cm",
            "không",
            "https://example.com/2d",
            "https://example.com/feedback",
        ],
        headers=headers,
    )

    response = _client().post(
        "/api/v1/products/import",
        files={
            "file": (
                "products.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    assert response.json()["imported_count"] == 1
    assert "shortDescription" not in response.json()["fields"]
    markdown = output_path.read_text(encoding="utf-8")
    assert "- **SKU**: S123" in markdown
    assert "- **Tên**: Sản phẩm test" in markdown
    assert "- **Giá**: 100000" in markdown
    assert "- **Mô tả**: " in markdown
    assert "- **Ảnh Lookbook**: https://example.com/lookbook" in markdown
    assert "- **Size**: " in markdown
    assert "- **Chiều dài**: 120cm" in markdown
    assert "- **Phụ kiện tặng kèm**: không" in markdown
    assert "- **Ảnh Cận chất**: https://example.com/2d" in markdown
    assert "- **Ảnh feedback**: https://example.com/feedback" in markdown
    assert "shortDescription" not in markdown
    assert "Không lưu mô tả ngắn" not in markdown
    assert "Không lưu cột thừa" not in markdown


def test_product_import_api_uses_hyperlink_target_instead_of_display_text(
    monkeypatch,
    tmp_path,
):
    output_path = tmp_path / "product.md"
    monkeypatch.setattr(product_import_service, "PRODUCT_MARKDOWN_PATH", output_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(product_import_service.PRODUCT_EXCEL_HEADERS)
    worksheet.append(
        [
            "S999",
            "Sản phẩm có link",
            "",
            "",
            "catalouge- Khánh Linh - Google Drive",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    lookbook_cell = worksheet.cell(
        row=2,
        column=product_import_service.PRODUCT_EXCEL_HEADERS.index("Ảnh Lookbook") + 1,
    )
    lookbook_cell.hyperlink = "https://drive.google.com/drive/folders/lookbook"
    lookbook_cell.style = "Hyperlink"

    response = _client().post(
        "/api/v1/products/import",
        files={
            "file": (
                "products.xlsx",
                _save_workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201
    markdown = output_path.read_text(encoding="utf-8")
    assert (
        "- **Ảnh Lookbook**: https://drive.google.com/drive/folders/lookbook"
        in markdown
    )
    assert "catalouge- Khánh Linh - Google Drive" not in markdown


def test_product_import_api_rejects_non_xlsx_file(monkeypatch, tmp_path):
    monkeypatch.setattr(
        product_import_service,
        "PRODUCT_MARKDOWN_PATH",
        tmp_path / "product.md",
    )

    response = _client().post(
        "/api/v1/products/import",
        files={"file": ("products.csv", b"Code,Name", "text/csv")},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == ErrorCode.INVALID_INPUT_DATA.value
