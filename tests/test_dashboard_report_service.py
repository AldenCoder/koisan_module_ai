import asyncio
from io import BytesIO
from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook

from app.api.dependencies.time import VN_TZ
from app.services import dashboard_report_service as service


class _FakeAggregateCursor:
    def __init__(self, rows):
        self._rows = rows

    async def to_list(self):
        return self._rows


def test_normalize_date_range_uses_full_vn_day_for_date_inputs():
    start, query_end, display_end = service._normalize_date_range(
        from_date="2026-06-01",
        to_date="2026-06-26",
    )

    assert start == datetime(2026, 6, 1, 0, 0, 0, tzinfo=VN_TZ)
    assert query_end == datetime(2026, 6, 27, 0, 0, 0, tzinfo=VN_TZ)
    assert display_end == datetime(2026, 6, 26, 23, 59, 59, 999999, tzinfo=VN_TZ)


def test_normalize_date_range_rejects_invalid_range():
    with pytest.raises(ValueError, match="from_date_must_be_before_to_date"):
        service._normalize_date_range(
            from_date="2026-06-26",
            to_date="2026-06-01",
        )


def test_normalize_date_range_rejects_too_large_range():
    with pytest.raises(ValueError, match="date_range_too_large"):
        service._normalize_date_range(
            from_date="2025-01-01",
            to_date="2026-12-31",
        )


def test_base_message_match_applies_filters():
    start = datetime(2026, 6, 1, tzinfo=VN_TZ)
    end = datetime(2026, 6, 2, tzinfo=VN_TZ)

    result = service._base_message_match(
        start=start,
        end=end,
        page_id="page-1",
        thread_type="comment",
        role="user",
    )

    assert result == {
        "created_at": {"$gte": start, "$lt": end},
        "meta.page_id": "page-1",
        "meta.message_type": "COMMENT",
        "role": "user",
    }


@pytest.mark.parametrize(
    "content",
    [
        "https://content.pancake.vn/2-2606/2026/6/25/sample.jpg",
        " https://content.pancake.vn/2-2606/2026/6/25/sample.PNG ",
        "https://content.pancake.vn/2-2606/2026/6/25/sample.webp?token=abc",
    ],
)
def test_content_classifier_detects_image_urls(content):
    assert service._is_image_content(content) is True
    assert service._is_text_content(content) is False


@pytest.mark.parametrize(
    "content",
    [
        "Mình muốn mẫu này",
        "https://example.com/product",
        "",
        None,
    ],
)
def test_content_classifier_detects_text_content(content):
    assert service._is_image_content(content) is False
    assert service._is_text_content(content) is bool(str(content or "").strip())


def test_message_projection_classifies_image_by_content_url_only():
    projection = service._message_projection_stage()["$project"]

    assert projection["has_image"] == {
        "$regexMatch": {
            "input": {"$trim": {"input": {"$ifNull": ["$content", ""]}}},
            "regex": service._IMAGE_CONTENT_URL_PATTERN,
            "options": "i",
        }
    }
    assert "meta.image" not in str(projection)


def test_support_reason_prioritizes_status_over_pause():
    now = datetime(2026, 6, 26, 10, 0, tzinfo=VN_TZ)

    assert service._support_reason({"status": "apilimit"}, now=now) == "apilimit"
    assert service._support_reason({"status": "handover"}, now=now) == "handover"
    assert (
        service._support_reason(
            {"status": "new", "bot_paused_until": now + timedelta(minutes=5)},
            now=now,
        )
        == "bot_paused"
    )


def test_get_dashboard_report_service_aggregates_report(monkeypatch):
    aggregate_calls = []
    updated_at = datetime(2026, 6, 26, 4, 0, 0)
    paused_until = datetime(2026, 6, 26, 4, 30, 0)

    message_rows = [
        {
            "summary": [
                {
                    "total_messages": 12,
                    "text_messages": 8,
                    "image_messages": 3,
                    "user_messages": 6,
                    "staff_messages": 2,
                    "bot_messages": 4,
                }
            ],
            "by_day": [
                {
                    "_id": "2026-06-26",
                    "total": 12,
                    "text": 8,
                    "image": 3,
                    "user": 6,
                    "staff": 2,
                    "bot": 4,
                }
            ],
        }
    ]
    conversation_status_rows = [
        {"_id": "new", "count": 5},
        {"_id": "handover", "count": 2},
        {"_id": "order_pending", "count": 1},
    ]
    needs_support_rows = [
        {
            "items": [
                {
                    "_id": "conv-support",
                    "customer_name": "Nguyen Van A",
                    "customer_id": "customer-1",
                    "status": "handover",
                    "pancake_page_id": "page-1",
                    "pancake_conversation_id": "pc-1",
                    "pancake_info_url": "https://pancake.vn/page-1?c_id=pc-1",
                    "order_note": None,
                    "bot_paused_until": paused_until,
                    "updated_at": updated_at,
                    "message_count": 7,
                }
            ],
            "total": [{"count": 2}],
        }
    ]
    order_rows = [
        {
            "items": [
                {
                    "_id": "conv-order",
                    "customer_name": "Tran Thi B",
                    "customer_id": "customer-2",
                    "status": "order_pending",
                    "pancake_page_id": "page-1",
                    "pancake_conversation_id": "pc-2",
                    "pancake_info_url": "https://pancake.vn/page-1?c_id=pc-2",
                    "order_note": "1. [10:05] Khach dat hang",
                    "updated_at": updated_at,
                    "message_count": 5,
                }
            ],
            "total": [{"count": 1}],
        }
    ]
    conversation_sequences = [conversation_status_rows, needs_support_rows, order_rows]

    def fake_message_aggregate(pipeline):
        aggregate_calls.append(("message", pipeline))
        return _FakeAggregateCursor(message_rows)

    def fake_conversation_aggregate(pipeline):
        aggregate_calls.append(("conversation", pipeline))
        return _FakeAggregateCursor(conversation_sequences.pop(0))

    monkeypatch.setattr(service.Message, "aggregate", fake_message_aggregate)
    monkeypatch.setattr(service.Conversation, "aggregate", fake_conversation_aggregate)
    monkeypatch.setattr(
        service,
        "now_vn",
        lambda: datetime(2026, 6, 26, 10, 0, 0, tzinfo=VN_TZ),
    )

    result = asyncio.run(
        service.get_dashboard_report_service(
            from_date="2026-06-26",
            to_date="2026-06-26",
            page_id="page-1",
            thread_type="inbox",
            role="user",
            alert_limit=10,
        )
    )

    assert result["summary"]["total_messages"] == 12
    assert result["summary"]["text_messages"] == 8
    assert result["summary"]["image_messages"] == 3
    assert result["summary"]["user_messages"] == 6
    assert result["summary"]["total_conversations"] == 8
    assert result["summary"]["handover_conversations"] == 2
    assert result["summary"]["order_pending_conversations"] == 1
    assert result["summary"]["needs_support_count"] == 2
    assert result["summary"]["order_alert_count"] == 1
    assert result["messages_by_day"] == [
        {
            "date": "2026-06-26",
            "total": 12,
            "text": 8,
            "image": 3,
            "user": 6,
            "staff": 2,
            "bot": 4,
        }
    ]
    assert result["alerts"]["needs_support"][0]["reason"] == "handover"
    assert result["alerts"]["needs_support"][0]["message_count"] == 7
    assert result["alerts"]["orders"][0]["order_note"] == "1. [10:05] Khach dat hang"
    assert result["filters"]["thread_type"] == "inbox"
    assert result["filters"]["role"] == "user"

    message_pipeline = aggregate_calls[0][1]
    assert message_pipeline[0]["$match"]["meta.page_id"] == "page-1"
    assert message_pipeline[0]["$match"]["meta.message_type"] == "INBOX"
    assert message_pipeline[0]["$match"]["role"] == "user"


def test_get_dashboard_report_service_returns_empty_payload(monkeypatch):
    conversation_sequences = [
        [],
        [{"items": [], "total": []}],
        [{"items": [], "total": []}],
    ]

    monkeypatch.setattr(
        service.Message,
        "aggregate",
        lambda pipeline: _FakeAggregateCursor([{"summary": [], "by_day": []}]),
    )
    monkeypatch.setattr(
        service.Conversation,
        "aggregate",
        lambda pipeline: _FakeAggregateCursor(conversation_sequences.pop(0)),
    )

    result = asyncio.run(
        service.get_dashboard_report_service(
            from_date="2026-06-26",
            to_date="2026-06-26",
        )
    )

    assert result["summary"]["total_messages"] == 0
    assert result["summary"]["text_messages"] == 0
    assert result["summary"]["image_messages"] == 0
    assert result["summary"]["total_conversations"] == 0
    assert result["messages_by_day"] == []
    assert result["alerts"] == {"needs_support": [], "orders": []}


def test_list_dashboard_report_page_ids_service_merges_conversations_and_messages(monkeypatch):
    first = datetime(2026, 6, 25, 10, 0, tzinfo=VN_TZ)
    second = datetime(2026, 6, 26, 10, 0, tzinfo=VN_TZ)
    aggregate_calls = []

    def fake_conversation_aggregate(pipeline):
        aggregate_calls.append(("conversation", pipeline))
        return _FakeAggregateCursor(
            [
                {
                    "_id": "page-1",
                    "conversation_count": 2,
                    "latest_activity_at": first,
                },
                {
                    "_id": "page-2",
                    "conversation_count": 1,
                    "latest_activity_at": second,
                },
            ]
        )

    def fake_message_aggregate(pipeline):
        aggregate_calls.append(("message", pipeline))
        return _FakeAggregateCursor(
            [
                {
                    "_id": "page-1",
                    "message_count": 7,
                    "latest_activity_at": second,
                },
                {
                    "_id": "page-3",
                    "message_count": 4,
                    "latest_activity_at": first,
                },
            ]
        )

    monkeypatch.setattr(service.Conversation, "aggregate", fake_conversation_aggregate)
    monkeypatch.setattr(service.Message, "aggregate", fake_message_aggregate)

    result = asyncio.run(service.list_dashboard_report_page_ids_service())

    assert result == {
        "items": [
            {
                "page_id": "page-1",
                "conversation_count": 2,
                "message_count": 7,
                "latest_activity_at": second,
            },
            {
                "page_id": "page-2",
                "conversation_count": 1,
                "message_count": 0,
                "latest_activity_at": second,
            },
            {
                "page_id": "page-3",
                "conversation_count": 0,
                "message_count": 4,
                "latest_activity_at": first,
            },
        ]
    }
    assert aggregate_calls[0][1][0]["$match"] == {
        "pancake_page_id": {"$exists": True, "$nin": [None, ""]},
        "is_active": True,
    }
    assert aggregate_calls[1][1][0]["$match"] == {
        "meta.page_id": {"$exists": True, "$nin": [None, ""]}
    }


def test_get_dashboard_report_service_rejects_invalid_filters():
    with pytest.raises(ValueError, match="invalid_thread_type"):
        asyncio.run(
            service.get_dashboard_report_service(
                from_date="2026-06-26",
                to_date="2026-06-26",
                thread_type="dm",
            )
        )


def test_collect_dashboard_report_export_details_returns_all_rows_without_limit(monkeypatch):
    now = datetime(2026, 6, 26, 10, 0, tzinfo=VN_TZ)
    aggregate_calls = []

    def fake_message_aggregate(pipeline):
        aggregate_calls.append(("message", pipeline))
        return _FakeAggregateCursor(
            [
                {
                    "_id": "msg-1",
                    "conversation_id": "conv-1",
                    "message_mid": "mid-1",
                    "role": "user",
                    "content": "Xin chào",
                    "meta": {
                        "page_id": "page-1",
                        "message_type": "INBOX",
                        "sender_id": "customer-1",
                    },
                    "created_at": now,
                    "updated_at": now,
                },
                {
                    "_id": "msg-2",
                    "conversation_id": "conv-1",
                    "message_mid": "mid-2",
                    "role": "staff",
                    "content": "https://cdn.example.com/photo.jpg",
                    "meta": {"page_id": "page-1"},
                    "created_at": now,
                    "updated_at": now,
                },
            ]
        )

    conversation_sequences = [
        [
            {
                "_id": "conv-new",
                "status": "new",
                "is_active": True,
                "message_count": 2,
                "created_at": now,
                "updated_at": now,
            }
        ],
        [
            {
                "_id": "conv-support",
                "status": "handover",
                "is_active": True,
                "message_count": 5,
                "created_at": now,
                "updated_at": now,
            }
        ],
        [
            {
                "_id": "conv-order",
                "status": "order_pending",
                "order_note": "Khách đặt hàng",
                "is_active": True,
                "message_count": 4,
                "created_at": now,
                "updated_at": now,
            }
        ],
    ]

    def fake_conversation_aggregate(pipeline):
        aggregate_calls.append(("conversation", pipeline))
        return _FakeAggregateCursor(conversation_sequences.pop(0))

    monkeypatch.setattr(service.Message, "aggregate", fake_message_aggregate)
    monkeypatch.setattr(service.Conversation, "aggregate", fake_conversation_aggregate)
    monkeypatch.setattr(service, "now_vn", lambda: now)

    result = asyncio.run(
        service._collect_dashboard_report_export_details(
            from_date="2026-06-26",
            to_date="2026-06-26",
            page_id="page-1",
            thread_type="inbox",
            role=None,
            include_inactive=False,
        )
    )

    assert len(result["messages"]) == 2
    assert result["messages"][0]["content_type"] == "text"
    assert result["messages"][1]["content_type"] == "image"
    assert len(result["new_conversations"]) == 1
    assert result["needs_support"][0]["support_reason"] == "handover"
    assert len(result["orders"]) == 1
    assert aggregate_calls[0][1][0]["$match"]["meta.page_id"] == "page-1"
    assert aggregate_calls[0][1][0]["$match"]["meta.message_type"] == "INBOX"
    assert all(
        "$limit" not in stage
        for _, pipeline in aggregate_calls
        for stage in pipeline
    )


def _export_report_payload():
    now = datetime(2026, 6, 26, 10, 0, tzinfo=VN_TZ)
    message_base = {
        "conversation_id": "conv-support",
        "page_id": "page-1",
        "thread_type": "INBOX",
        "source": "pancake_webhook",
        "sender_id": "customer-1",
        "sender_name": "Nguyễn Văn A",
        "pancake_conversation_id": "pc-1",
        "created_at": now,
        "updated_at": now,
    }
    conversation_base = {
        "channel": "Koisan",
        "customer_name": "Nguyễn Văn A",
        "customer_id": "customer-1",
        "pancake_page_id": "page-1",
        "pancake_thread_type": "inbox",
        "pancake_info_url": "https://pancake.vn/page-1?c_id=pc-1",
        "is_active": True,
        "summaries": ["Khách đang cần tư vấn"],
        "fb_ai_initialized": True,
        "fb_ai_initialized_at": now,
        "bot_paused_until": now,
        "bot_paused_at": now,
        "bot_paused_reason": "handover",
        "bot_paused_by": "agent-1",
        "created_at": now,
        "updated_at": now,
    }
    return {
        "filters": {
            "from_date": now,
            "to_date": now,
            "page_id": "page-1",
            "thread_type": "inbox",
            "role": None,
            "include_inactive": False,
            "alert_limit": 20,
        },
        "summary": {
            "total_messages": 4,
            "text_messages": 2,
            "image_messages": 1,
            "user_messages": 1,
            "staff_messages": 1,
            "bot_messages": 2,
            "total_conversations": 2,
            "new_conversations": 1,
            "confirmed_conversations": 0,
            "handover_conversations": 0,
            "apilimit_conversations": 0,
            "order_pending_conversations": 1,
            "needs_support_count": 1,
            "order_alert_count": 1,
        },
        "messages_by_day": [
            {
                "date": "2026-06-26",
                "total": 4,
                "text": 2,
                "image": 1,
                "user": 1,
                "staff": 1,
                "bot": 2,
            }
        ],
        "conversation_status": [
            {"status": "new", "count": 1},
            {"status": "handover", "count": 0},
            {"status": "order_pending", "count": 1},
        ],
        "alerts": {
            "needs_support": [
                {
                    "conversation_id": "conv-support",
                    "customer_name": "Nguyen Van A",
                    "customer_id": "customer-1",
                    "status": "handover",
                    "reason": "handover",
                    "pancake_page_id": "page-1",
                    "pancake_conversation_id": "pc-1",
                    "pancake_info_url": "https://pancake.vn/page-1?c_id=pc-1",
                    "order_note": None,
                    "bot_paused_until": now,
                    "updated_at": now,
                    "message_count": 7,
                }
            ],
            "orders": [
                {
                    "conversation_id": "conv-order",
                    "customer_name": "Tran Thi B",
                    "customer_id": "customer-2",
                    "status": "order_pending",
                    "pancake_page_id": "page-1",
                    "pancake_conversation_id": "pc-2",
                    "pancake_info_url": "https://pancake.vn/page-1?c_id=pc-2",
                    "order_note": "1. [10:05] Khach dat hang",
                    "updated_at": now,
                    "message_count": 5,
                }
            ],
        },
        "export_details": {
            "messages": [
                {
                    **message_base,
                    "message_id": "msg-user",
                    "message_mid": "mid-user",
                    "content_type": "text",
                    "role": "user",
                    "content": "=2+2",
                    "image_urls": "",
                    "meta": {"ghi_chu": "Tiếng Việt đầy đủ"},
                },
                {
                    **message_base,
                    "message_id": "msg-staff",
                    "message_mid": "mid-staff",
                    "content_type": "image",
                    "role": "staff",
                    "content": "https://cdn.example.com/anh-san-pham.jpg",
                    "image_urls": "https://cdn.example.com/anh-san-pham.jpg",
                    "meta": {"image_urls": ["https://cdn.example.com/anh-san-pham.jpg"]},
                },
                {
                    **message_base,
                    "message_id": "msg-bot-text",
                    "message_mid": None,
                    "content_type": "text",
                    "role": "bot",
                    "content": "Agent đã phản hồi",
                    "image_urls": "",
                    "meta": {"action": "reply"},
                },
                {
                    **message_base,
                    "message_id": "msg-bot-empty",
                    "message_mid": None,
                    "content_type": "other",
                    "role": "bot",
                    "content": "",
                    "image_urls": "",
                    "meta": {},
                },
            ],
            "new_conversations": [
                {
                    **conversation_base,
                    "conversation_id": "conv-new",
                    "pancake_conversation_id": "pc-new",
                    "status": "new",
                    "order_note": None,
                    "message_count": 2,
                },
                {
                    **conversation_base,
                    "conversation_id": "conv-order",
                    "pancake_conversation_id": "pc-order",
                    "status": "order_pending",
                    "order_note": "Khách đặt 01 sản phẩm",
                    "message_count": 2,
                },
            ],
            "needs_support": [
                {
                    **conversation_base,
                    "conversation_id": "conv-support",
                    "pancake_conversation_id": "pc-support",
                    "status": "handover",
                    "support_reason": "handover",
                    "order_note": None,
                    "message_count": 3,
                }
            ],
            "orders": [
                {
                    **conversation_base,
                    "conversation_id": "conv-order",
                    "pancake_conversation_id": "pc-order",
                    "status": "order_pending",
                    "order_note": "Khách đặt 01 sản phẩm",
                    "message_count": 2,
                }
            ],
        },
    }


def test_build_dashboard_report_workbook_contains_expected_sheets_and_headers():
    workbook = service.build_dashboard_report_workbook(_export_report_payload())

    assert workbook.sheetnames == [
        "Tổng quan",
        "Theo ngày",
        "Trạng thái hội thoại",
        "Tất cả tin nhắn",
        "Tin nhắn văn bản",
        "Tin nhắn hình ảnh",
        "Tin nhắn khách",
        "Tin nhắn nhân viên",
        "Tin nhắn Agent",
        "Hội thoại tạo mới",
        "Cần hỗ trợ",
        "Cảnh báo đơn hàng",
    ]
    summary_sheet = workbook["Tổng quan"]
    assert summary_sheet["A2"].value == (
        "File đối soát chi tiết, mỗi chỉ số có tab chi tiết tương ứng."
    )
    assert [summary_sheet.cell(row, 1).value for row in range(14, 19)] == [
        "Từ ngày",
        "Đến ngày",
        "Trang",
        "Loại hội thoại",
        "Người gửi",
    ]
    message_sheet = workbook["Tất cả tin nhắn"]
    assert [cell.value for cell in message_sheet[5]][:7] == [
        "STT",
        "Mã tin nhắn nền tảng",
        "ID hội thoại hệ thống",
        "Loại nội dung",
        "Người gửi",
        "Nội dung tin nhắn",
        "URL hình ảnh",
    ]
    assert "ID tin nhắn hệ thống" not in [cell.value for cell in message_sheet[5]]
    assert message_sheet.freeze_panes == "A6"
    assert message_sheet["A3"].value == "Số bản ghi: 4"
    assert message_sheet["F6"].value == "=2+2"
    assert message_sheet["F6"].data_type == "s"
    assert "Tiếng Việt đầy đủ" in message_sheet["P6"].value
    assert message_sheet["C6"].hyperlink.target == (
        "https://koisan-dashboard.vercel.app/quan-ly-lich-su/conv-support"
    )
    assert workbook["Tin nhắn văn bản"]["A3"].value == "Số bản ghi: 2"
    assert workbook["Tin nhắn hình ảnh"]["A3"].value == "Số bản ghi: 1"
    assert workbook["Tin nhắn khách"]["A3"].value == "Số bản ghi: 1"
    assert workbook["Tin nhắn nhân viên"]["A3"].value == "Số bản ghi: 1"
    assert workbook["Tin nhắn Agent"]["A3"].value == "Số bản ghi: 2"
    assert workbook["Hội thoại tạo mới"]["A3"].value == "Số bản ghi: 2"
    assert workbook["Hội thoại tạo mới"]["B6"].hyperlink.target == (
        "https://koisan-dashboard.vercel.app/quan-ly-lich-su/conv-new"
    )
    assert workbook["Cần hỗ trợ"]["E6"].value == "Chuyển cho nhân viên"
    assert workbook["Cảnh báo đơn hàng"]["K6"].hyperlink.target.startswith("https://")


def test_export_dashboard_report_excel_service_returns_xlsx_bytes(monkeypatch):
    async def fake_report_service(**kwargs):
        payload = _export_report_payload()
        payload["summary"]["total_messages"] = 999
        return payload

    async def fake_export_details(**kwargs):
        return _export_report_payload()["export_details"]

    monkeypatch.setattr(service, "get_dashboard_report_service", fake_report_service)
    monkeypatch.setattr(
        service,
        "_collect_dashboard_report_export_details",
        fake_export_details,
    )

    content, filename = asyncio.run(
        service.export_dashboard_report_excel_service(
            from_date="2026-06-26",
            to_date="2026-06-26",
        )
    )

    assert filename == "bao-cao-du-lieu-2026-06-26-den-2026-06-26.xlsx"
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == [
        "Tổng quan",
        "Theo ngày",
        "Trạng thái hội thoại",
        "Tất cả tin nhắn",
        "Tin nhắn văn bản",
        "Tin nhắn hình ảnh",
        "Tin nhắn khách",
        "Tin nhắn nhân viên",
        "Tin nhắn Agent",
        "Hội thoại tạo mới",
        "Cần hỗ trợ",
        "Cảnh báo đơn hàng",
    ]
    assert workbook["Theo ngày"]["B6"].value == "2026-06-26"
    assert workbook["Tất cả tin nhắn"]["A3"].value == "Số bản ghi: 4"
    assert workbook["Tổng quan"]["A5"].value == 4
    assert workbook["Tất cả tin nhắn"]["C6"].hyperlink.target == (
        "https://koisan-dashboard.vercel.app/quan-ly-lich-su/conv-support"
    )
